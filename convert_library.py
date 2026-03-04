#!/usr/bin/env python3
"""
Library CSV Converter
=====================
Converts the scanning app's xlsx/csv export to the format
expected by the hedgehog-library CSV importer.

Usage:
    python convert_library.py librar.xlsx
    python convert_library.py librar.xlsx --output ready_to_import.csv
"""

import sys
import csv
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


def convert(input_path: str, output_path: str):
    path = Path(input_path)

    if path.suffix.lower() in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    elif path.suffix.lower() == ".csv":
        import csv as _csv
        with open(input_path, encoding="utf-8-sig") as f:
            reader = _csv.DictReader(f)
            headers = reader.fieldnames
            rows = [tuple(row[h] for h in headers) for row in reader]
    else:
        print(f"Unsupported format: {path.suffix}")
        sys.exit(1)

    def fmt_date(val):
        if val is None:
            return ""
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        return str(val).strip()

    def fmt_isbn(val):
        if val is None:
            return ""
        # Remove any decimal point from integer representation
        s = str(val).strip()
        if s.endswith(".0"):
            s = s[:-2]
        return s

    def fmt_str(val):
        if val is None:
            return ""
        return str(val).strip()

    # Map source columns to hedgehog importer columns
    HEDGEHOG_COLS = [
        "title", "creators", "first_name", "last_name",
        "ean_isbn13", "upc_isbn10", "description", "publisher",
        "publish_date", "tags", "notes", "status", "added", "length"
    ]

    h = {v: i for i, v in enumerate(headers)}

    def get(row, col, default=""):
        idx = h.get(col)
        if idx is None:
            return default
        v = row[idx]
        return v if v is not None else default

    output_rows = []
    skipped = 0

    for row in rows:
        item_type = fmt_str(get(row, "item_type")).lower()
        if item_type and item_type != "book":
            skipped += 1
            continue

        output_rows.append({
            "title":        fmt_str(get(row, "title")) or "Unknown Title",
            "creators":     fmt_str(get(row, "creators")),
            "first_name":   fmt_str(get(row, "first_name")),
            "last_name":    fmt_str(get(row, "last_name")),
            "ean_isbn13":   fmt_isbn(get(row, "ean_isbn13")),
            "upc_isbn10":   fmt_isbn(get(row, "upc_isbn10")),
            "description":  fmt_str(get(row, "description")),
            "publisher":    fmt_str(get(row, "publisher")),
            "publish_date": fmt_date(get(row, "publish_date")),
            "tags":         fmt_str(get(row, "tags")),
            "notes":        fmt_str(get(row, "notes")),
            "status":       fmt_str(get(row, "status")),
            "added":        fmt_date(get(row, "added")),
            "length":       fmt_str(get(row, "length")),
        })

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEDGEHOG_COLS)
        writer.writeheader()
        writer.writerows(output_rows)

    print(f"Converted {len(output_rows)} books -> {output_path}")
    if skipped:
        print(f"Skipped {skipped} non-book items (games, discs, etc.)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Path to .xlsx or .csv export from scanning app")
    parser.add_argument("--output", default="hedgehog_import.csv",
                        help="Output CSV path (default: hedgehog_import.csv)")
    args = parser.parse_args()
    convert(args.input, args.output)
